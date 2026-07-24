from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Calendário Editorial"

HEADERS = [
    "Número do post", "Data", "Horário", "Tema", "Formato", "Pilar",
    "Objetivo", "Descrição da imagem", "Texto da imagem", "Legenda",
]

rows = [
[1, "08/07/2026", "12h00", "Registro do encontro (a partir do sábado 04/07)", "Reel", "Movimento em ação",
 "Consciência + prova social — mostrar que o encontro acontece de verdade, com frequência",
 "Vídeo de registro real do encontro de sábado (04/07) — checklist de 4 planos: close em mãos/instrumento, reação genuína, momento de dificuldade real, plano geral por último. Luz natural do espaço, sem filtro pesado, áudio real do encontro (sem música adicionada).",
 "Overlay de abertura (opcional): \"Foi assim no último encontro.\" | CTA final: \"A agenda fica nos destaques.\"",
 "Foi assim no último encontro: quem toca pela primeira vez do lado de quem já não sabe mais viver sem isso. A gente constrói o Bumba Meu Boi junto — ritmo, instrumento, escuta. Toda semana rola. Quer ver de perto? A agenda fica nos destaques. 🥁\n\n#BantuKatu #SeteLagoas #CulturaSeteLagoas #Percussão #BumbaMeuBoi #TamboresDoBantuKatu"],

[2, "10/07/2026", "19h00", "Ciclo de Vivência: Bumba Meu Boi", "Carrossel", "As frentes",
 "Consideração informativa — explicar o que é o ciclo, sem CTA de participação (pedido do cliente, 2026-07-09)",
 "Capa: foto real de encontro com véu escuro + tipografia forte. Slide 2: card educativo, textura ou foto de instrumento tradicional. Slide 3: elemento gráfico de linha do tempo (jun–jul–ago). Slide 4: foto real de grupo tocando junto, sem pose. Slide 5: mesmo estilo gráfico do slide 3, fecha o carrossel.",
 "Capa: \"O ciclo que está rolando agora\"\nSlide 2 — \"Esse ritmo tem raiz\": Bumba Meu Boi é uma celebração tradicional brasileira — nasceu da mistura de indígenas, africanos e portugueses. Tem canto, dança, teatro e percussão. Aqui no Bantu-Katu, a gente estuda os ritmos do boi através dos instrumentos que tocamos.\nSlide 3 — \"De junho a agosto\": O ciclo começou em junho e vai até agosto — encontros toda semana, dentro do Tambores do Bantu-Katu. É um recorte temático: cada ciclo trabalha uma tradição diferente.\nSlide 4 — \"Quem faz parte\": No ciclo tem gente que já toca há anos e gente que nunca pegou um instrumento antes — cada um no seu ritmo, dentro do mesmo grupo.\nSlide 5 — \"Como funciona\": Os encontros acontecem toda semana. O ciclo atual é o Bumba Meu Boi — depois dele, vem outro tema.",
 "O Bumba Meu Boi é o ciclo que o Tambores do Bantu-Katu está trabalhando desde junho, até agosto. Ritmos, instrumentos, história — tudo dentro do que já rola por aqui toda semana.\n\n#BantuKatu #CulturaBrasileira #BumbaMeuBoi #Percussão #SeteLagoas #CulturaSeteLagoas #MovimentoCultural #TamboresDoBantuKatu"],

[3, "13/07/2026", "12h00", "Registro do encontro — adaptação do sotaque de matraca (a partir do sábado 11/07)", "Reel", "Movimento em ação",
 "Consciência + prova social — mostrar a adaptação do repertório tradicional em tempo real",
 "Reel de registro real — close na mão tocando matraca, corte para a mesma mão (ou outra) tocando o instrumento do movimento (tumbadora/tambor) com o mesmo padrão rítmico. Contraste visual da transição, luz natural, som real, sem música de fundo.",
 "Overlay: \"A matraca tem seu jeito\" → \"Vira tumbadora / Vira tambor\" → CTA final: \"A agenda fica nos destaques 🥁\"",
 "A matraca tem seu jeito — e aqui ela ganha outro corpo. No encontro de sábado, a gente trouxe o sotaque da matraca pro instrumento do movimento: mesma pulsação, som diferente. É assim que a tradição segue viva no Bantu-Katu. A agenda fica nos destaques. 🥁\n\n#BantuKatu #SeteLagoas #CulturaAfroBrasileira #Percussão #BumbaMeuBoi #MovimentoCultural"],

[4, "15/07/2026", "12h00", "Registro real", "Foto", "Movimento em ação",
 "Consciência + prova social",
 "Foto real do encontro — checklist de 4 planos (mãos/instrumento, reação genuína, dificuldade real, plano geral). Luz natural, sem pose, enquadramento horizontal.",
 "Sem overlay definido (foto única + legenda)",
 "Foi assim no sábado: quem chegou pela primeira vez do lado de quem já não sabe mais viver sem isso. Tem matracas encontrando pandeirões, tem erro que vira acerto dois tempos depois, tem o som do boi tomando forma. É o Ciclo do Bumba Meu Boi rolando dentro do Tambores do Bantu-Katu. Qualquer pessoa toca? Toca. A coisa é séria — mas o movimento vem até você. A agenda fica nos destaques. 🥁\n\n#BantuKatu #SeteLagoas #CulturaAfroBrasileira #BumbaMeuBoi #Percussão #CulturaSeteLagoas #MovimentoCultural #MinasGerais"],

[5, "17/07/2026", "19h00", "Toadas de Bumba Meu Boi — Bruno canta \"Venha meu vaqueiro\"", "Reel", "Cultura viva",
 "Consciência ampliada + autoridade cultural",
 "Reel de registro real: o Bruno cantando/tocando a toada \"Venha meu vaqueiro que o boi vai sair...\". Close nas mãos no instrumento e no Bruno cantando, luz natural, áudio real (a voz e o instrumento são a trilha).",
 "Sem overlay definido — vídeo cru + legenda na caption. Letra real da toada: \"Venha meu vaqueiro que o boi vai sair, quero ver meu boi janeiro brincar no terreiro... bumba bumba bumba meu boi\"",
 "Essa é uma das toadas que o Bruno canta no Bantu-Katu: \"Venha meu vaqueiro que o boi vai sair, quero ver meu boi janeiro brincar no terreiro...\"\n\nAs toadas são as cantigas do Bumba Meu Boi — não é uma música só, é um repertório inteiro. Cada toada tem seu momento dentro da brincadeira: tem toada de chamada, de celebração, de despedida. É por elas que a história inteira se conta, do começo ao fim.\n\nNo Ciclo do Bumba Meu Boi, a gente tá aprendendo esse repertório aos poucos — cantando, tocando, guardando cada toada nova.\n\nSalva pra não perder — tem mais toada vindo por aí!\n\n#BantuKatu #CulturaAfroBrasileira #BumbaMeuBoi #Toadas #CulturaPopular #SeteLagoas #MovimentoCultural"],

[6, "20/07/2026", "12h00", "Registro do encontro — parte de uma toada (a partir do sábado 18/07)", "Reel", "Movimento em ação",
 "Consciência + prova social",
 "Reel de registro real — close nas mãos/instrumento tocando o fragmento da toada, groove já em andamento desde o início. Som real, sem edição, sem música de fundo.",
 "Overlay opcional: \"Toada do Bumba Meu Boi / Encontro de sábado\" | CTA final: \"A agenda fica nos destaques\"",
 "Foi assim no último sábado: a toada tocada no encontro de percussão do Bantu-Katu. O ritmo tem nome, e a história tem raiz. O Tambores do Bantu-Katu segue toda semana. Quer ver de perto? A agenda fica nos destaques. O movimento vem até você. 🥁\n\n#BantuKatu #SeteLagoas #CulturaAfroBrasileira #Percussão #BumbaMeuBoi #MovimentoCultural"],

[7, "22/07/2026", "19h00", "Contrate o Bantu-Katu", "Imagem única", "As frentes",
 "Conversão comercial — gerar contato pra contratação de apresentações/eventos (CTA comercial liberado em 2026-07-09)",
 "Foto real de apresentação/ensaio — instrumento em primeiro plano nítido, instrumentistas em movimento, participantes desfocados ao fundo. Luz natural, mood quente/terroso.",
 "Sem headline própria — comunicação concentrada na legenda",
 "O Bantu-Katu toca pra quem quer ouvir. Contrata aí.\n\nA gente leva percussão, ritmo e cultura afro-brasileira pro seu evento, festa, celebração, projeto — qualquer espaço que caiba som de verdade e gente que sabe o que tá fazendo. O movimento vem até você.\n\nQuer conversar? Chama na DM ou manda um e-mail: [e-mail a confirmar com Bruno]. Ou vem nos Stories — lá a gente posta os próximos encontros.\n\n#BantuKatu #SeteLagoas #CulturaSeteLagoas #CulturaAfroBrasileira #Percussão #EventosCulturais #CulturaPopular #MovimentoCultural"],

[8, "24/07/2026", "19h00", "O tambor: o que é, como soa, que ritmos toca no ciclo do Bumba Meu Boi", "Carrossel + áudio", "Cultura viva — Sonoridades Bantu-Katu",
 "Consciência ampliada + autoridade cultural",
 "Capa com tipografia forte. Slides seguintes: close real no tambor (foto ou frame de vídeo). Áudio real do tambor tocando embutido no slide 3.",
 "Capa: \"Tambor — o coração ritmado do movimento\"\nSlide 2 — \"Esse instrumento tem nome e origem\": O tambor é um cilindro de madeira e couro que traz consigo séculos de diálogo entre a tradição africana e a música popular brasileira. No Bantu-Katu, ele não apenas marca o ritmo — ele é presença.\nSlide 3 — \"Escuta como soa aqui\" [áudio real do tambor]: No ciclo do Bumba Meu Boi que rola com a gente agora, o tambor é quem sustenta a base.\nSlide 4 — \"O papel dele no nosso ritmo\": No Bumba Meu Boi, o tambor toca continuum — aquela pulsação que não sai do lugar, que deixa tudo marcado.",
 "Salva pra não perder — e vem acompanhar de perto como a gente constrói tudo isso junto.\n\n#BantuKatu #Tambor #CulturaAfroBrasileira #Percussão #CulturaPopular #SeteLagoas"],

[9, "27/07/2026", "12h00", "Fragmento de uma música do ciclo (a partir do encontro de 25/07)", "Vídeo/foto de registro", "Movimento em ação — Repertório Bantu-Katu",
 "Consciência + prova social — mostrar o repertório musical sem entregar tudo de uma vez",
 "Vídeo real de um trecho musical identificável do encontro de 25/07 — close no instrumento protagonista da música, luz natural, áudio limpo (prioridade — é o que carrega o post). Música ainda a definir com o Bruno.",
 "Sem overlay definido — vídeo cru + legenda na caption",
 "Primeira parte de uma música que está no nosso repertório agora. O que você tá ouvindo é do Ciclo do Bumba Meu Boi — uma tradição que a gente tá aprendendo, tocando e reconstruindo juntos.\n\nEsse fragmento é só o começo. A música inteira fica pro próximo mês.\n\nQuer ouvir o resto? Fica de olho por aqui. 🥁\n\n#RepertórioBantuKatu #BumbaOBoi #CicloPercussivo #BantuKatu #CulturaAfroBrasileira #SeteLagoas #CulturaPopular"],

[10, "29/07/2026", "12h00", "Registros do Cazuá do Rei Divino (participação a confirmar)", "Carrossel de imagens reais", "Movimento em ação",
 "A definir — pedido do cliente pra desenvolver depois",
 "A definir — falta confirmar participação e material real do evento",
 "A definir",
 "A definir — não desenvolver ainda (falta contexto real sobre o que é o Cazuá do Rei Divino e se a participação do Bantu-Katu se confirma)"],

[11, "03/08/2026", "12h00", "Registro do encontro — parte da explicação do Bruno (a partir do sábado 01/08)", "Reel", "Movimento em ação",
 "Consciência + prova social",
 "Reel de registro real do Bruno explicando algo do ciclo/ritmo pro grupo, seguido do grupo tocando junto. Luz natural, som real. ATENÇÃO: usar a fala real que o Bruno disser na captura — não roteirizar antecipadamente o que ele vai dizer.",
 "Sem overlay fixo definido — usar apenas se necessário pra legibilidade",
 "Foi assim no último encontro: o Bruno explicando como o Bumba Meu Boi conversa através dos instrumentos — e todo mundo tocando junto pra entender na prática. O Bantu-Katu rola todo sábado, aberto pra quem quer chegar. A agenda fica nos destaques. O movimento vem até você. 🥁\n\n#BantuKatu #SeteLagoas #CulturaAfroBrasileira #Percussão #BumbaMeuBoi #MovimentoCultural"],

[12, "05/08/2026", "19h00", "Parcerias: Boi Batukado (colaboração com o Quintal Boi da Manta)", "Carrossel", "As frentes",
 "A definir — pedido do cliente pra desenvolver depois",
 "A definir",
 "A definir",
 "A definir — não desenvolver ainda (falta contexto real sobre a parceria)"],

[13, "12/08/2026", "19h00", "Dicionário Bantu-Katu (palavra a definir)", "Imagem única", "Cultura viva",
 "Consciência ampliada + autoridade cultural",
 "Card educativo, mesmo sistema visual do Dicionário Bantukatu (fundo terroso, tipografia display forte). Estrutura pronta, falta só a palavra.",
 "[TERMO A DEFINIR] — estrutura: De onde vem (origem/tradição/significado) → Como soa → No Bantu-Katu (instrumento do movimento que assume o papel)",
 "[TERMO A DEFINIR] — Palavra de raiz [origem], que significa [significado]. De onde vem: [contexto tradicional]. Como soa: [descrição sonora]. No Bantu-Katu: [instrumento do movimento que assume o papel]. Salva esse dicionário e volta sempre que quiser lembrar de onde vem a música que a gente toca por aqui. 🥁\n\n#BantuKatu #DicionárioBantuKatu #CulturaAfroBrasileira #CulturaPopular #SeteLagoas — falta a palavra escolhida pelo Bruno pra fechar o texto"],

[14, "14/08/2026", "19h00", "A tumbadora: o que é, como soa, que ritmos toca no ciclo do Bumba Meu Boi", "Carrossel + áudio", "Cultura viva — Sonoridades Bantu-Katu",
 "Consciência ampliada + autoridade cultural",
 "Capa com tipografia forte. Close real na tumbadora (foto ou frame de vídeo). Áudio real da tumbadora embutido no slide 2.",
 "Capa: \"Esse instrumento vem do coração\"\nSlide 1 — \"Tumbadora: tambor de raiz cubana\": Ela vem de Cuba, mas toca em ritmos que atravessam a gente toda. Dois tons — grave e agudo — que conversam entre si. No Bantu-Katu, ela carrega o coração do Bumba Meu Boi.\nSlide 2 — \"Escuta só esse som\" [áudio real da tumbadora]: Grave fundo, agudo que corta. Ela toca sozinha, mas responde pro tambor.\nSlide 3 — \"No ciclo do Bumba Meu Boi\": A tumbadora marca o pulso, aquele que faz o corpo virar. Quando ela entra junto do tambor, é aí que o ritmo ganha corpo — é ela que puxa quem quer dançar.",
 "Tem muito som por trás de cada encontro do Bantu-Katu. Cada instrumento tem nome, origem e função — e cada um deles toca na gente de um jeito diferente. Salva essa série pra não perder. 🥁\n\n#BantuKatu #SonoridadesBantuKatu #BumbaMeuBoi #Tumbadora #CulturaAfroBrasileira #CulturaPopular #Percussão #SeteLagoas"],

[15, "17/08/2026", "12h00", "Registro do encontro (a partir do sábado 15/08)", "Carrossel de fotos reais", "Movimento em ação",
 "Consciência + prova social",
 "Carrossel de fotos reais do encontro — mãos/instrumento, clima humano do encontro, plano geral do grupo tocando junto. Luz natural, sem pose.",
 "Sem overlay definido — cards de foto + legenda na caption",
 "Foi assim no sábado: quem chegou pela primeira vez do lado de quem já não sabe mais viver sem isso. O corpo aprende antes da mente — a gente entra escutando, sai tocando, vai embora dançando.\n\nCiclo do Bumba Meu Boi: o repertório já tem forma. Matracas, pandeirões, tambor, tumbadora — cada som em seu lugar.\n\nTem gente que chegou pro primeiro Bumba e já tá diferente de quem toca desde junho. Não é nível — é enraizamento. O movimento vem até você.\n\nA agenda fica nos destaques. Chama a gente se ficar com dúvida.\n\n#SeteLagoas #CulturaSeteLagoas #MinasGerais #BumbaMeuBoi #Percussão #CulturaAfroBrasileira #BantuKatu"],

[16, "19/08/2026", "19h00", "Divulgação Encontro de Bois", "Reforço", "As frentes",
 "A definir — pedido do cliente pra desenvolver depois",
 "A definir",
 "A definir",
 "A definir — não desenvolver ainda (falta contexto real sobre o Encontro de Bois)"],

[17, "21/08/2026", "19h00", "Divulgação Encontro de Bois", "Reforço", "As frentes",
 "A definir — pedido do cliente pra desenvolver depois",
 "A definir",
 "A definir",
 "A definir — não desenvolver ainda (falta contexto real sobre o Encontro de Bois)"],

[18, "24/08/2026", "12h00", "Uma música do ciclo, no arranjo completo (a partir do encontro de 22/08)", "Vídeo/foto de registro", "Movimento em ação — Repertório Bantu-Katu",
 "Consciência + prova social — segunda entrega da série, mostrar o repertório completo",
 "Vídeo real de uma música completa do ciclo — instrumento principal em primeiro plano, demais instrumentos desfocados ao fundo, luz natural. Música a definir (diferente da usada em julho).",
 "Sem overlay definido — vídeo cru + legenda na caption",
 "Assim é quando o arranjo fica completo. [NOME DA MÚSICA], do ciclo que a gente está construindo agora — tambor, tumbadora, tamborim, efeitos. A roda fica diferente quando cada instrumento encontra seu lugar.\n\nO Bantu-Katu é uma rede de projetos que trabalha ritmo de verdade. Toda semana a gente se encontra pra aprofundar esse repertório.\n\nQuer ver como fica quando tudo junto? Vem acompanhando por aqui. A agenda dos encontros fica nos destaques.\n\n#Percussão #BantuKatu #CulturaAfroBrasileira #SeteLagoas #CulturaPopular #MovimentoCultural #MinasGerais #CulturaSeteLagoas"],

[19, "26/08/2026", "19h00", "Aprendizados do Ciclo do Boi", "Carrossel", "As frentes",
 "Consideração — retrospectiva do ciclo, sem tom de despedida",
 "Carrossel com 1-2 fotos reais do ciclo (mãos tocando, close em instrumento, rosto em escuta) do acervo já existente — evitar pose de \"encerramento\".",
 "Capa: \"O Boi nos ensinou a tocar juntos\"\nSlide 2 — \"Ritmo é construção coletiva\": Não é cada um tocando o seu — é o tambor conversando com a tumbadora, o tamborim abrindo espaço, os efeitos respirando no meio.\nSlide 3 — \"A tradição pulsa quando é viva\": Aprendemos que Bumba não é coisa de livro: é movimento, é corpo, é encontro que muda a cada semana conforme quem chega.\nSlide 4 — \"Crescemos tocando, não planejando\": Três meses é pouco pra dominar um ciclo — e isso é o ponto. O Bantu-Katu cresceu porque a gente tocou junto, errou junto, ajustou junto.\nSlide 5: O próximo ciclo já está sendo construído.",
 "O Boi nos trouxe pra cá — agora a gente segue ouvindo, tocando, aprendendo. Acompanha o que vem aí? A agenda fica nos destaques.\n\n#BantuKatu #CulturaSeteLagoas #MovimentoCultural"],

[20, "31/08/2026", "20h00", "O que vem por aí", "Teaser", "Bora junto",
 "Engajamento ativo — gerar curiosidade e comentário sobre o próximo ciclo, sem revelar tema se ainda não estiver definido",
 "Composição em camadas: close macro de um instrumento em ação (registro real já existente) + textura sutil (madeira/palha) + fundo cor sólida terrosa/quente. Tipografia display forte.",
 "\"O próximo já vem sendo construído.\" (headline a validar com design)",
 "O Bumba Meu Boi do Bantu-Katu marca seus ritmos de despedida — e marca também o começo de algo novo. Um ciclo é só isso: volta que gira e volta a girar.\n\nDo bumba ao próximo. As mãos já estão preparando, a cabeça esquentando, o repertório sendo construído. Tem gente do Bantu-Katu pensando nele, tocando nele desde agora — porque o movimento não funciona como interrupção e retomada, funciona como fluxo.\n\nQuem quer saber o que vem por aí manda uma mensagem. O movimento vem até você. 🥁\n\n#BantuKatu #SeteLagoas #CulturaAfroBrasileira #CulturaPopular #MovimentoCultural"],
]

FONT_NAME = "Arial"
header_fill = PatternFill("solid", start_color="8B6F47", end_color="8B6F47")
header_font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
body_font = Font(name=FONT_NAME, size=10)
pending_font = Font(name=FONT_NAME, size=10, italic=True, color="9C5B00")
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws.append(HEADERS)
for col in range(1, len(HEADERS) + 1):
    c = ws.cell(row=1, column=col)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
ws.row_dimensions[1].height = 22

for r in rows:
    ws.append(r)

for row_idx in range(2, len(rows) + 2):
    is_pending = ws.cell(row=row_idx, column=7).value == "A definir — pedido do cliente pra desenvolver depois"
    for col_idx in range(1, len(HEADERS) + 1):
        c = ws.cell(row=row_idx, column=col_idx)
        c.font = pending_font if is_pending else body_font
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.border = border
    ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="top")
    ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center", vertical="top")
    ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="center", vertical="top")

widths = {1: 8, 2: 13, 3: 12, 4: 34, 5: 16, 6: 22, 7: 30, 8: 40, 9: 42, 10: 55}
for col_idx, w in widths.items():
    ws.column_dimensions[get_column_letter(col_idx)].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:J{len(rows) + 1}"

wb.save("clientes/bantu-katu/estrategia_2026-07-03/calendario-editorial-julho-agosto-2026.xlsx")
print("Salvo.")
